from __future__ import annotations

import io
import re
import shutil
import subprocess
import tempfile
from pathlib import Path

import pikepdf
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connection
from django.test import TestCase, TransactionTestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook
from reportlab.pdfgen import canvas

from PIL import Image, ImageChops

from editor.forms import DocumentTemplateForm
from editor.models import DocumentTemplate, TemplateField
from editor.services import read_page_geometry, update_template_pdf_metadata
from fonts.models import FontAsset

from . import services
from .forms import GenerationJobForm
from .models import GenerationItem, GenerationJob
from .services import _merge_overlay, format_field_value, process_job

TEST_MEDIA_ROOT = tempfile.mkdtemp(prefix="sistema_vetorial_jobs_tests_")


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT)
class JobTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="jobs-user", password="senha123")
        font_path = Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
        cls.font = FontAsset.objects.create(
            user=cls.user,
            name="Dejavu Sans",
            family="Dejavu Sans",
            variant="Regular",
            file=SimpleUploadedFile(
                "DejaVuSans.ttf", font_path.read_bytes(), content_type="font/ttf"
            ),
        )

    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)

    def _build_pdf_upload(self) -> SimpleUploadedFile:
        temp_dir = Path(tempfile.mkdtemp(prefix="jobs-pdf-"))
        pdf_path = temp_dir / "background.pdf"
        pdf_canvas = canvas.Canvas(str(pdf_path), pagesize=(400, 120))
        pdf_canvas.drawString(24, 100, "Pagina 1")
        pdf_canvas.showPage()
        pdf_canvas.save()
        content = pdf_path.read_bytes()
        shutil.rmtree(temp_dir, ignore_errors=True)
        return SimpleUploadedFile("background.pdf", content, content_type="application/pdf")

    def _build_excel_upload(self, rows=None) -> SimpleUploadedFile:
        temp_dir = Path(tempfile.mkdtemp(prefix="jobs-xlsx-"))
        xlsx_path = temp_dir / "dados.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Nome", "Empresa"])
        for row in rows or [["Ana", "Empresa A"], ["Bruno", "Empresa B"]]:
            sheet.append(row)
        workbook.save(xlsx_path)
        content = xlsx_path.read_bytes()
        shutil.rmtree(temp_dir, ignore_errors=True)
        return SimpleUploadedFile(
            "dados.xlsx",
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _build_template(self):
        template = DocumentTemplate.objects.create(
            user=self.user,
            name="Faixa",
            slug="faixa",
            background_pdf=self._build_pdf_upload(),
        )
        update_template_pdf_metadata(template)
        TemplateField.objects.create(
            template=template,
            name="nome",
            excel_column="1",
            x=20,
            y=80,
            width=200,
            height=24,
            font=self.font,
            font_size=18,
            order_index=1,
            text_transform=TemplateField.TextTransform.TITLE_SMART,
            transform_exceptions="de, da, do, dos, e",
        )
        return template

    def test_format_field_value_supports_title_transform(self):
        template = self._build_template()
        field = template.fields.get()
        formatted = format_field_value(field, "  joao da silva  ")
        self.assertEqual(formatted, "Joao da Silva")

    def test_job_form_validates_column_numbers(self):
        template = self._build_template()
        template.fields.update(excel_column="3")
        form = GenerationJobForm(
            data={"name": "Job", "template": template.pk},
            files={"source_excel": self._build_excel_upload()},
            user=self.user,
        )
        self.assertFalse(form.is_valid())
        self.assertIn("colunas numéricas", form.non_field_errors()[0])

    def test_process_job_uses_numeric_columns(self):
        template = self._build_template()
        template.fields.update(
            border_enabled=True,
            border_color="#112233",
            border_size_ratio=0.08,
            border_opacity=0.75,
            border_blur=0.4,
        )
        job = GenerationJob.objects.create(
            user=self.user,
            template=template,
            name="Job",
            source_excel=self._build_excel_upload(),
            kind=GenerationJob.Kind.PREVIEW,
            status=GenerationJob.Status.QUEUED,
        )

        process_job(job)
        job.refresh_from_db()

        self.assertEqual(job.success_rows, 2)
        self.assertEqual(job.items.count(), 2)
        first_item = job.items.order_by("id").first()
        with (
            first_item.output_pdf.open("rb") as output_file,
            pikepdf.Pdf.open(output_file) as output_pdf,
        ):
            media_box = [float(value) for value in output_pdf.pages[0].MediaBox]
        self.assertEqual(media_box[2] - media_box[0], 400.0)
        self.assertEqual(media_box[3] - media_box[1], 120.0)

    def test_job_delete_marks_inactive(self):
        template = self._build_template()
        job = GenerationJob.objects.create(
            user=self.user,
            template=template,
            name="Job Inativo",
            source_excel=self._build_excel_upload(),
        )
        self.client.force_login(self.user)

        response = self.client.post(reverse("jobs:delete", kwargs={"pk": job.pk}))

        self.assertEqual(response.status_code, 302)
        job.refresh_from_db()
        self.assertFalse(job.is_active)

    def test_launch_api_generates_preview_files(self):
        template = self._build_template()
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("jobs:launch"),
            {
                "template": template.pk,
                "kind": GenerationJob.Kind.PREVIEW,
                "name": "Lote da bancada",
                "source_excel": self._build_excel_upload(),
            },
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        job = GenerationJob.objects.get(pk=payload["id"])
        self.assertEqual(job.kind, GenerationJob.Kind.PREVIEW)
        self.assertEqual(payload["status_url"], reverse("jobs:status", kwargs={"pk": job.pk}))

        # CELERY_TASK_ALWAYS_EAGER=1 nos testes: o job roda dentro do request.
        job.refresh_from_db()
        self.assertEqual(job.status, GenerationJob.Status.COMPLETED)
        self.assertEqual(job.success_rows, 2)

        status = self.client.get(payload["status_url"]).json()
        self.assertEqual(status["status"], "completed")
        self.assertEqual(status["processed_rows"], 2)
        self.assertEqual(len(status["items"]), 2)
        self.assertTrue(all(item["output_url"] for item in status["items"]))

    def test_launch_api_full_kind_exposes_zip(self):
        template = self._build_template()
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("jobs:launch"),
            {
                "template": template.pk,
                "kind": GenerationJob.Kind.FULL,
                "source_excel": self._build_excel_upload(),
            },
        )

        self.assertEqual(response.status_code, 201)
        status = self.client.get(response.json()["status_url"]).json()
        self.assertEqual(status["kind"], "full")
        self.assertTrue(status["zip_url"])
        zip_response = self.client.get(status["zip_url"])
        self.assertEqual(zip_response.status_code, 200)

    def test_launch_api_names_the_job_when_the_name_is_blank(self):
        template = self._build_template()
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("jobs:launch"),
            {
                "template": template.pk,
                "kind": GenerationJob.Kind.PREVIEW,
                "name": "   ",
                "source_excel": self._build_excel_upload(),
            },
        )

        job = GenerationJob.objects.get(pk=response.json()["id"])
        self.assertTrue(job.name.startswith(template.name))

    def test_launch_api_rejects_missing_excel(self):
        template = self._build_template()
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("jobs:launch"),
            {"template": template.pk, "kind": GenerationJob.Kind.PREVIEW},
        )

        self.assertEqual(response.status_code, 400)
        self.assertTrue(response.json()["error"])
        self.assertEqual(GenerationJob.objects.count(), 0)

    def test_launch_api_rejects_non_excel_upload(self):
        template = self._build_template()
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("jobs:launch"),
            {
                "template": template.pk,
                "kind": GenerationJob.Kind.PREVIEW,
                "source_excel": SimpleUploadedFile(
                    "dados.csv", b"Nome,Empresa", content_type="text/csv"
                ),
            },
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("Excel", response.json()["error"])

    def test_launch_api_rejects_invalid_kind(self):
        template = self._build_template()
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("jobs:launch"),
            {
                "template": template.pk,
                "kind": "amostra",
                "source_excel": self._build_excel_upload(),
            },
        )

        self.assertEqual(response.status_code, 400)

    def test_launch_api_rejects_template_without_fields(self):
        template = self._build_template()
        template.fields.all().delete()
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("jobs:launch"),
            {
                "template": template.pk,
                "kind": GenerationJob.Kind.PREVIEW,
                "source_excel": self._build_excel_upload(),
            },
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("campo", response.json()["error"])

    def test_launch_api_ignores_templates_of_other_users(self):
        template = self._build_template()
        other_user = get_user_model().objects.create_user(
            username="jobs-outro", password="senha123"
        )
        self.client.force_login(other_user)

        response = self.client.post(
            reverse("jobs:launch"),
            {
                "template": template.pk,
                "kind": GenerationJob.Kind.PREVIEW,
                "source_excel": self._build_excel_upload(),
            },
        )

        self.assertEqual(response.status_code, 404)

    def test_launch_api_requires_login(self):
        template = self._build_template()

        response = self.client.post(
            reverse("jobs:launch"),
            {
                "template": template.pk,
                "kind": GenerationJob.Kind.PREVIEW,
                "source_excel": self._build_excel_upload(),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn("/login/", response["Location"])

    def test_status_endpoint_hides_jobs_of_other_users(self):
        template = self._build_template()
        job = GenerationJob.objects.create(
            user=self.user,
            template=template,
            name="Job",
            source_excel=self._build_excel_upload(),
        )
        other_user = get_user_model().objects.create_user(
            username="jobs-bisbilhoteiro", password="senha123"
        )
        self.client.force_login(other_user)

        response = self.client.get(reverse("jobs:status", kwargs={"pk": job.pk}))

        self.assertEqual(response.status_code, 404)

    def test_status_endpoint_reports_the_failure_reason(self):
        template = self._build_template()
        template.fields.update(excel_column="9")
        job = GenerationJob.objects.create(
            user=self.user,
            template=template,
            name="Job com coluna faltando",
            source_excel=self._build_excel_upload(),
            kind=GenerationJob.Kind.PREVIEW,
            status=GenerationJob.Status.QUEUED,
        )
        process_job(job)
        self.client.force_login(self.user)

        payload = self.client.get(reverse("jobs:status", kwargs={"pk": job.pk})).json()

        self.assertEqual(payload["status"], "failed")
        self.assertIn("Coluna 9", payload["last_error"])
        self.assertEqual(payload["zip_url"], "")


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT)
class JobProgressVisibilityTests(TransactionTestCase):
    """O card de geração só mostra progresso porque `process_job` grava fora de
    uma transação; envolver o job inteiro em `atomic` deixaria o polling parado
    até o commit final."""

    def test_process_job_saves_progress_outside_a_transaction(self):
        user = get_user_model().objects.create_user(username="jobs-progress", password="senha123")
        font_path = Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
        font = FontAsset.objects.create(
            user=user,
            name="Dejavu Sans",
            family="Dejavu Sans",
            variant="Regular",
            file=SimpleUploadedFile(
                "DejaVuSans.ttf", font_path.read_bytes(), content_type="font/ttf"
            ),
        )
        pdf_dir = Path(tempfile.mkdtemp(prefix="jobs-progress-pdf-"))
        pdf_path = pdf_dir / "background.pdf"
        pdf_canvas = canvas.Canvas(str(pdf_path), pagesize=(400, 120))
        pdf_canvas.showPage()
        pdf_canvas.save()
        template = DocumentTemplate.objects.create(
            user=user,
            name="Faixa",
            slug="faixa-progresso",
            background_pdf=SimpleUploadedFile(
                "background.pdf", pdf_path.read_bytes(), content_type="application/pdf"
            ),
        )
        shutil.rmtree(pdf_dir, ignore_errors=True)
        update_template_pdf_metadata(template)
        TemplateField.objects.create(
            template=template,
            name="nome",
            excel_column="1",
            x=20,
            y=80,
            width=200,
            height=24,
            font=font,
            font_size=18,
            order_index=1,
        )

        xlsx_dir = Path(tempfile.mkdtemp(prefix="jobs-progress-xlsx-"))
        xlsx_path = xlsx_dir / "dados.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Nome"])
        sheet.append(["Ana"])
        sheet.append(["Bruno"])
        workbook.save(xlsx_path)
        job = GenerationJob.objects.create(
            user=user,
            template=template,
            name="Job",
            source_excel=SimpleUploadedFile("dados.xlsx", xlsx_path.read_bytes()),
            kind=GenerationJob.Kind.PREVIEW,
            status=GenerationJob.Status.QUEUED,
        )
        shutil.rmtree(xlsx_dir, ignore_errors=True)

        seen_inside_atomic = []
        original = services._create_item_pdf

        def spy(*args, **kwargs):
            seen_inside_atomic.append(connection.in_atomic_block)
            return original(*args, **kwargs)

        services._create_item_pdf = spy
        try:
            process_job(job)
        finally:
            services._create_item_pdf = original

        self.assertEqual(seen_inside_atomic, [False, False])
        job.refresh_from_db()
        self.assertEqual(job.processed_rows, 2)
        job.items.all().delete()
        job.delete()


def _rasterize(pdf_bytes: bytes, resolution: int = 72):
    """Rasteriza a página (o pdftoppm aplica /Rotate e recorta pela CropBox,
    igual ao preview do editor) e devolve uma imagem em tons de cinza."""
    temp_dir = Path(tempfile.mkdtemp(prefix="jobs-raster-"))
    try:
        pdf_path = temp_dir / "page.pdf"
        pdf_path.write_bytes(pdf_bytes)
        subprocess.run(
            [
                "pdftoppm",
                "-png",
                "-cropbox",
                "-r",
                str(resolution),
                "-gray",
                str(pdf_path),
                str(temp_dir / "p"),
            ],
            check=True,
            capture_output=True,
        )
        rendered = sorted(temp_dir.glob("p-*.png"))[0]
        with Image.open(rendered) as image:
            return image.convert("L").copy()
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def _ink_bbox(pdf_bytes: bytes, resolution: int = 72):
    gray = _rasterize(pdf_bytes, resolution)
    return ImageChops.invert(gray).getbbox(), gray.size


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT)
class OutputFidelityTests(TestCase):
    """O PDF de saída precisa ser o PDF de entrada com curvas por cima: mesmas
    caixas de página, mesma imagem de fundo byte a byte e texto vetorizado."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="fidelity", password="senha123")
        font_path = Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
        cls.font = FontAsset.objects.create(
            user=cls.user,
            name="Dejavu Sans",
            family="Dejavu Sans",
            variant="Regular",
            file=SimpleUploadedFile(
                "DejaVuSans.ttf", font_path.read_bytes(), content_type="font/ttf"
            ),
        )

    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)

    def _background(self, *, size=(400, 120), rotation=0, cropbox=None) -> SimpleUploadedFile:
        # Página em branco montada pelo pikepdf: sem conteúdo nenhum, para que
        # qualquer operador no PDF de saída venha comprovadamente da geração.
        with pikepdf.Pdf.new() as pdf:
            page = pdf.add_blank_page(page_size=size)
            if rotation:
                page.Rotate = rotation
            if cropbox:
                page.CropBox = pikepdf.Array([float(value) for value in cropbox])
            output = io.BytesIO()
            pdf.save(output)
        return SimpleUploadedFile("fundo.pdf", output.getvalue(), content_type="application/pdf")

    def _template(self, background, *, x=20, y=20, font_size=18, width=200):
        template = DocumentTemplate.objects.create(
            user=self.user,
            name=f"Fidelidade {DocumentTemplate.objects.count()}",
            slug=f"fidelidade-{DocumentTemplate.objects.count()}",
            background_pdf=background,
        )
        update_template_pdf_metadata(template)
        TemplateField.objects.create(
            template=template,
            name="nome",
            excel_column="1",
            x=x,
            y=y,
            width=width,
            height=font_size * 1.5,
            font=self.font,
            font_size=font_size,
            order_index=1,
        )
        return template

    def _excel(self, value="Ana") -> SimpleUploadedFile:
        buffer = io.BytesIO()
        workbook = Workbook()
        workbook.active.append(["Nome"])
        workbook.active.append([value])
        workbook.save(buffer)
        return SimpleUploadedFile("dados.xlsx", buffer.getvalue())

    def _generate(self, template, value="Ana") -> bytes:
        job = GenerationJob.objects.create(
            user=self.user,
            template=template,
            name="Job",
            source_excel=self._excel(value),
            kind=GenerationJob.Kind.PREVIEW,
            status=GenerationJob.Status.QUEUED,
        )
        process_job(job)
        job.refresh_from_db()
        self.assertEqual(job.status, GenerationJob.Status.COMPLETED, job.last_error)
        with job.items.get().output_pdf.open("rb") as generated:
            return generated.read()

    def test_output_carries_only_vector_paths(self):
        template = self._template(self._background())

        output = self._generate(template, "Ação Wq")

        with pikepdf.Pdf.open(io.BytesIO(output)) as pdf:
            content = b" ".join(pikepdf.Page(page).obj.Contents.read_bytes() for page in pdf.pages)
            forms = [
                obj
                for obj in pdf.objects
                if isinstance(obj, pikepdf.Stream) and obj.get("/Subtype") == pikepdf.Name.Form
            ]
            content += b" ".join(form.read_bytes() for form in forms)
            embedded_fonts = [
                key
                for obj in pdf.objects
                if isinstance(obj, pikepdf.Dictionary | pikepdf.Stream)
                for key in obj.keys()
                if str(key).startswith("/FontFile")
            ]

        self.assertFalse(embedded_fonts, "o PDF vetorial não deve embutir fontes")
        self.assertIsNone(
            re.search(rb"(^|\s)(BT|Tj|TJ)(\s|$)", content),
            "o texto deve sair como curva, não como operador de texto",
        )
        # m/l/c/h: os operadores de contorno que sobraram no lugar do texto.
        self.assertGreater(len(re.findall(rb"\bc[\s\n]", content)), 10)

    def test_output_keeps_the_background_image_untouched(self):
        source = Image.new("RGB", (240, 90), (250, 240, 220))
        for x in range(0, 240, 6):
            for y in range(90):
                source.putpixel((x, y), (10, 40, 90))
        png_buffer = io.BytesIO()
        source.save(png_buffer, format="PNG")
        form = DocumentTemplateForm(
            data={"name": "Fundo PNG", "description": ""},
            files={
                "background_pdf": SimpleUploadedFile("fundo.png", png_buffer.getvalue()),
            },
        )
        form.instance.user = self.user
        self.assertTrue(form.is_valid(), form.errors)
        template = form.save()
        update_template_pdf_metadata(template)
        TemplateField.objects.create(
            template=template,
            name="nome",
            excel_column="1",
            x=10,
            y=10,
            width=150,
            height=20,
            font=self.font,
            font_size=14,
            order_index=1,
        )

        with pikepdf.Pdf.open(template.background_pdf.path) as background:
            background_image = background.pages[0].Resources.XObject.Im0
            background_bytes = background_image.read_raw_bytes()
            background_dims = (int(background_image.Width), int(background_image.Height))
            background_filter = str(background_image.get("/Filter"))
            background_box = [float(value) for value in background.pages[0].MediaBox]

        output = self._generate(template)

        with pikepdf.Pdf.open(io.BytesIO(output)) as pdf:
            page = pdf.pages[0]
            images = [
                obj
                for obj in page.Resources.XObject.values()
                if obj.get("/Subtype") == pikepdf.Name.Image
            ]
            self.assertEqual(len(images), 1)
            self.assertEqual(images[0].read_raw_bytes(), background_bytes)
            self.assertEqual((int(images[0].Width), int(images[0].Height)), background_dims)
            self.assertEqual(str(images[0].get("/Filter")), background_filter)
            self.assertEqual([float(value) for value in page.MediaBox], background_box)

    def test_output_keeps_the_page_boxes_of_the_background(self):
        template = self._template(self._background(cropbox=[20, 10, 380, 110]))

        output = self._generate(template)

        with pikepdf.Pdf.open(io.BytesIO(output)) as pdf:
            page = pdf.pages[0]
            self.assertEqual([float(v) for v in page.MediaBox], [0.0, 0.0, 400.0, 120.0])
            self.assertEqual([float(v) for v in page.CropBox], [20.0, 10.0, 380.0, 110.0])

    def test_template_measures_the_visible_box_not_the_mediabox(self):
        cropped = self._template(self._background(cropbox=[20, 10, 380, 110]))
        rotated = self._template(self._background(rotation=90))

        self.assertEqual((float(cropped.page_width), float(cropped.page_height)), (360.0, 100.0))
        self.assertEqual((float(rotated.page_width), float(rotated.page_height)), (120.0, 400.0))

    def test_text_lands_inside_the_cropbox(self):
        template = self._template(self._background(cropbox=[20, 10, 380, 110]), x=5, y=5)

        bbox, size = _ink_bbox(self._generate(template))

        self.assertIsNotNone(bbox, "o texto precisa aparecer dentro da área visível")
        self.assertEqual(size, (360, 100))
        # Campo em (5, 5) a partir do topo-esquerda da CropBox.
        self.assertLess(bbox[0], 40)
        self.assertLess(bbox[1], 40)

    def test_text_stays_upright_on_a_rotated_background(self):
        template = self._template(self._background(rotation=90), x=10, y=10, width=90, font_size=14)

        bbox, size = _ink_bbox(self._generate(template))

        self.assertIsNotNone(bbox)
        # A página girada mede 120 x 400 na tela, e é nessa medida que o campo
        # foi posicionado.
        self.assertEqual(size, (120, 400))
        self.assertLess(bbox[0], 60, "o texto deve ficar na metade esquerda")
        self.assertLess(bbox[1], 120, "o texto deve ficar perto do topo")
        # Deitado, o texto sairia mais alto do que largo.
        self.assertGreater(bbox[2] - bbox[0], bbox[3] - bbox[1])

    def test_merge_refuses_a_background_that_changed_size(self):
        template = self._template(self._background())
        geometry = read_page_geometry(template.background_pdf.path)
        overlay_buffer = io.BytesIO()
        overlay_canvas = canvas.Canvas(overlay_buffer, pagesize=(geometry.width, geometry.height))
        overlay_canvas.showPage()
        overlay_canvas.save()
        wrong = geometry._replace(x1=geometry.x1 + 50)

        with self.assertRaises(ValueError) as raised:
            _merge_overlay(template.background_pdf.path, overlay_buffer.getvalue(), wrong)

        self.assertIn("Reenvie o fundo", str(raised.exception))


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT)
class JobDetailAndStatusWindowingTests(TestCase):
    """A tela do job e o polling de status precisam aguentar um lote grande:
    a tabela pagina, os chips filtram por estado e o JSON de status nunca
    devolve o lote inteiro de uma vez."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="jobs-windowing", password="senha123"
        )

    def _template(self) -> DocumentTemplate:
        template = DocumentTemplate.objects.create(
            user=self.user,
            name="Lote grande",
            slug=f"lote-grande-{DocumentTemplate.objects.count()}",
            background_pdf=self._build_pdf_upload(),
        )
        update_template_pdf_metadata(template)
        return template

    def _build_pdf_upload(self) -> SimpleUploadedFile:
        temp_dir = Path(tempfile.mkdtemp(prefix="jobs-window-pdf-"))
        pdf_path = temp_dir / "background.pdf"
        pdf_canvas = canvas.Canvas(str(pdf_path), pagesize=(400, 120))
        pdf_canvas.showPage()
        pdf_canvas.save()
        content = pdf_path.read_bytes()
        shutil.rmtree(temp_dir, ignore_errors=True)
        return SimpleUploadedFile("background.pdf", content, content_type="application/pdf")

    def _job_with_items(self, *, total: int, failed_rows: set[int] = frozenset()) -> GenerationJob:
        template = self._template()
        job = GenerationJob.objects.create(
            user=self.user,
            template=template,
            name="Lote grande",
            source_excel=self._build_excel_upload(rows=[["x", "y"]] * total),
            kind=GenerationJob.Kind.FULL,
            status=GenerationJob.Status.COMPLETED,
            total_rows=total,
            processed_rows=total,
            success_rows=total - len(failed_rows),
            failed_rows=len(failed_rows),
        )
        for row_number in range(1, total + 1):
            is_failed = row_number in failed_rows
            GenerationItem.objects.create(
                job=job,
                row_number=row_number,
                status=(
                    GenerationItem.Status.FAILED if is_failed else GenerationItem.Status.COMPLETED
                ),
                error_message="Falha de teste" if is_failed else "",
            )
        return job

    def _build_excel_upload(self, rows) -> SimpleUploadedFile:
        buffer = io.BytesIO()
        workbook = Workbook()
        workbook.active.append(["Nome", "Empresa"])
        for row in rows:
            workbook.active.append(row)
        workbook.save(buffer)
        return SimpleUploadedFile("dados.xlsx", buffer.getvalue())

    def test_detail_paginates_items_in_windows_of_25(self):
        job = self._job_with_items(total=40)
        self.client.force_login(self.user)

        first_page = self.client.get(reverse("jobs:detail", kwargs={"pk": job.pk}))
        second_page = self.client.get(reverse("jobs:detail", kwargs={"pk": job.pk}), {"page": 2})

        self.assertEqual(len(first_page.context["page_obj"].object_list), 25)
        self.assertEqual(len(second_page.context["page_obj"].object_list), 15)
        self.assertEqual(first_page.context["page_obj"].paginator.num_pages, 2)

    def test_detail_filters_items_by_estado(self):
        job = self._job_with_items(total=10, failed_rows={2, 5, 9})
        self.client.force_login(self.user)

        failed = self.client.get(
            reverse("jobs:detail", kwargs={"pk": job.pk}), {"estado": "failed"}
        )
        completed = self.client.get(
            reverse("jobs:detail", kwargs={"pk": job.pk}), {"estado": "completed"}
        )

        self.assertEqual(failed.context["page_obj"].paginator.count, 3)
        self.assertEqual(
            {item.row_number for item in failed.context["page_obj"].object_list}, {2, 5, 9}
        )
        self.assertEqual(completed.context["page_obj"].paginator.count, 7)

    def test_detail_ignores_an_unknown_estado(self):
        job = self._job_with_items(total=5)
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("jobs:detail", kwargs={"pk": job.pk}), {"estado": "sabotagem"}
        )

        self.assertEqual(response.context["estado"], "")
        self.assertEqual(response.context["page_obj"].paginator.count, 5)

    def test_detail_shows_chip_counts_and_pagination_controls(self):
        job = self._job_with_items(total=40, failed_rows={1, 2, 3})
        self.client.force_login(self.user)

        html = self.client.get(reverse("jobs:detail", kwargs={"pk": job.pk})).content.decode()

        self.assertIn('id="chip-count-all">40', html)
        self.assertIn('id="chip-count-completed">37', html)
        self.assertIn('id="chip-count-failed">3', html)
        self.assertIn("Página 1 de 2", html)
        self.assertIn('id="job-items-next"', html)
        self.assertIn('href="?page=2"', html)

    def test_status_endpoint_windows_items_by_default(self):
        job = self._job_with_items(total=40)
        self.client.force_login(self.user)

        payload = self.client.get(reverse("jobs:status", kwargs={"pk": job.pk})).json()

        self.assertEqual(len(payload["items"]), 25)
        self.assertEqual(payload["items_count"], 40)
        self.assertEqual(payload["items_num_pages"], 2)
        self.assertTrue(payload["items_has_next"])
        self.assertFalse(payload["items_has_previous"])
        self.assertEqual(payload["pending_rows"], 0)
        self.assertIn(f"/jobs/{job.pk}/", payload["detail_url"])

    def test_status_endpoint_respects_estado_and_page(self):
        job = self._job_with_items(total=40, failed_rows={4, 8})
        self.client.force_login(self.user)

        payload = self.client.get(
            reverse("jobs:status", kwargs={"pk": job.pk}), {"estado": "failed", "page": 1}
        ).json()

        self.assertEqual(payload["items_count"], 2)
        self.assertEqual({item["row_number"] for item in payload["items"]}, {4, 8})
        self.assertEqual(payload["items_num_pages"], 1)
        self.assertFalse(payload["items_has_next"])

    def test_status_endpoint_hides_other_users_jobs_even_with_filters(self):
        job = self._job_with_items(total=3)
        other = get_user_model().objects.create_user(username="jobs-window-outro", password="x")
        self.client.force_login(other)

        response = self.client.get(
            reverse("jobs:status", kwargs={"pk": job.pk}), {"estado": "failed"}
        )

        self.assertEqual(response.status_code, 404)

    def test_detail_page_uses_lote_vocabulary(self):
        job = self._job_with_items(total=3)
        self.client.force_login(self.user)

        html = self.client.get(reverse("jobs:detail", kwargs={"pk": job.pk})).content.decode()

        self.assertIn('<div class="eyebrow">Lote</div>', html)
        self.assertIn("Resumo do lote", html)
        self.assertNotIn(">Job<", html)


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT)
class PortugueseCopyTests(TestCase):
    """ "Job"/"Jobs" era o único vocabulário em inglês sobrando nas telas de
    lote — o app inteiro (inclusive o link "Arquivos" da topbar) já falava
    "lote"/"arquivos". Trocado por consistência com o resto do produto."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.user = get_user_model().objects.create_user(username="jobs-copy", password="senha123")

    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)

    def test_job_list_uses_lote_vocabulary(self):
        self.client.force_login(self.user)

        html = self.client.get(reverse("jobs:list")).content.decode()

        self.assertIn("Seus arquivos", html)
        self.assertIn("Novo lote", html)
        self.assertNotIn(">Jobs<", html)
        self.assertNotIn("Nenhum job encontrado", html)

    def test_job_create_form_uses_lote_vocabulary(self):
        self.client.force_login(self.user)

        html = self.client.get(reverse("jobs:create")).content.decode()

        self.assertIn("Novo lote", html)
        self.assertIn("Criar lote", html)
        self.assertIn("Nome do lote", html)
        self.assertNotIn("Nome do job", html)

    def test_job_model_and_app_are_named_in_portuguese(self):
        self.assertEqual(GenerationJob._meta.verbose_name, "lote")
        self.assertEqual(GenerationJob._meta.verbose_name_plural, "lotes")
        self.assertEqual(GenerationItem._meta.verbose_name, "item do lote")


class ColumnTemplateTests(TestCase):
    """`extract_column_refs`/`render_column_template` são a mini-sintaxe {N}
    compartilhada entre "colunas do Excel" de um campo (concatenar colunas com
    um conector) e o padrão de nome de arquivo do template."""

    def test_extract_column_refs_bare_digit(self):
        self.assertEqual(services.extract_column_refs("3"), [3])

    def test_extract_column_refs_multiple_braces(self):
        self.assertEqual(services.extract_column_refs("{1}_{2}"), [1, 2])

    def test_extract_column_refs_repeated_column(self):
        self.assertEqual(services.extract_column_refs("{3}-{3}"), [3, 3])

    def test_extract_column_refs_blank(self):
        self.assertEqual(services.extract_column_refs(""), [])
        self.assertEqual(services.extract_column_refs(None), [])

    def test_render_column_template_bare_digit(self):
        self.assertEqual(services.render_column_template("1", {"coluna_1": "Ana"}), "Ana")

    def test_render_column_template_concatenates_with_connector(self):
        payload = {"coluna_3": "Recife", "coluna_5": "PE"}
        self.assertEqual(services.render_column_template("{3}-{5}", payload), "Recife-PE")

    def test_render_column_template_missing_column_becomes_empty(self):
        self.assertEqual(services.render_column_template("{9}", {}), "")

    def test_render_column_template_blank_pattern(self):
        self.assertEqual(services.render_column_template("", {"coluna_1": "Ana"}), "")


class ResolveFieldRawValueTests(TestCase):
    """`resolve_field_raw_value` decide se a Transformação roda coluna a
    coluna (campo com 2+ colunas e `transform_columns` marcado) ou, como
    sempre foi, sobre o texto já unido (campo de 1 coluna, ou sem escolha)."""

    def _field(self, **overrides):
        defaults = {
            "excel_column": "1",
            "text_transform": TemplateField.TextTransform.UPPER,
            "transform_columns": [],
            "trim_whitespace": True,
        }
        defaults.update(overrides)
        return TemplateField(**defaults)

    def test_single_column_never_splits_transform(self):
        field = self._field(excel_column="1", transform_columns=[1])
        raw_value, transform_applied = services.resolve_field_raw_value(
            field, {"coluna_1": "recife"}
        )
        self.assertEqual(raw_value, "recife")
        self.assertFalse(transform_applied)

    def test_multi_column_without_selection_keeps_old_behavior(self):
        field = self._field(excel_column="{3}-{5}", transform_columns=[])
        raw_value, transform_applied = services.resolve_field_raw_value(
            field, {"coluna_3": "recife", "coluna_5": "pe"}
        )
        self.assertEqual(raw_value, "recife-pe")
        self.assertFalse(transform_applied)

    def test_multi_column_with_selection_transforms_only_selected_columns(self):
        field = self._field(excel_column="{3}-{5}", transform_columns=[5])
        raw_value, transform_applied = services.resolve_field_raw_value(
            field, {"coluna_3": "recife", "coluna_5": "pe"}
        )
        self.assertEqual(raw_value, "recife-PE")
        self.assertTrue(transform_applied)

    def test_format_field_value_skips_transform_when_already_applied(self):
        field = self._field(excel_column="{3}-{5}", transform_columns=[5])
        formatted = format_field_value(field, "recife-PE", transform_applied=True)
        self.assertEqual(formatted, "recife-PE")


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT)
class FilenamePatternTests(TestCase):
    """`DocumentTemplate.filename_pattern` decide o nome de cada PDF gerado; sem
    padrão, cai no nome antigo (lote + linha). Colisões (o mesmo nome vindo de
    duas linhas) ganham um sufixo para não se sobrescreverem dentro do ZIP."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="filenames-user", password="senha123"
        )
        font_path = Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
        cls.font = FontAsset.objects.create(
            user=cls.user,
            name="Dejavu Sans",
            family="Dejavu Sans",
            variant="Regular",
            file=SimpleUploadedFile(
                "DejaVuSans.ttf", font_path.read_bytes(), content_type="font/ttf"
            ),
        )

    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)

    def _build_pdf_upload(self) -> SimpleUploadedFile:
        buffer = io.BytesIO()
        pdf_canvas = canvas.Canvas(buffer, pagesize=(400, 120))
        pdf_canvas.showPage()
        pdf_canvas.save()
        return SimpleUploadedFile(
            "background.pdf", buffer.getvalue(), content_type="application/pdf"
        )

    def _build_excel_upload(self, rows) -> SimpleUploadedFile:
        temp_dir = Path(tempfile.mkdtemp(prefix="filenames-xlsx-"))
        xlsx_path = temp_dir / "dados.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Nome", "Empresa"])
        for row in rows:
            sheet.append(row)
        workbook.save(xlsx_path)
        content = xlsx_path.read_bytes()
        shutil.rmtree(temp_dir, ignore_errors=True)
        return SimpleUploadedFile(
            "dados.xlsx",
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _build_template(
        self,
        filename_pattern="",
        filename_space_mode=DocumentTemplate.FilenameSpaceMode.KEEP,
        filename_space_replacement="-",
        filename_strip_accents=False,
        filename_case=DocumentTemplate.FilenameCase.NONE,
        filename_case_columns=None,
    ) -> DocumentTemplate:
        slug_bits = (
            f"{filename_pattern or 'padrao'}-{filename_space_mode}"
            f"-{filename_strip_accents}-{filename_case}-{filename_case_columns}"
        )
        template = DocumentTemplate.objects.create(
            user=self.user,
            name="Crachá",
            slug=f"cracha-{re.sub(r'[^a-z0-9]+', '-', slug_bits.lower())}",
            background_pdf=self._build_pdf_upload(),
            filename_pattern=filename_pattern,
            filename_space_mode=filename_space_mode,
            filename_space_replacement=filename_space_replacement,
            filename_strip_accents=filename_strip_accents,
            filename_case=filename_case,
            filename_case_columns=filename_case_columns or [],
        )
        update_template_pdf_metadata(template)
        TemplateField.objects.create(
            template=template,
            name="nome",
            excel_column="1",
            x=20,
            y=80,
            width=200,
            height=24,
            font=self.font,
            font_size=18,
            order_index=1,
        )
        return template

    def test_sanitize_filename_component_strips_unsafe_chars(self):
        # Caractere proibido em nome de arquivo vira espaço mesmo no modo
        # padrão "manter espaços" — só o texto original é preservado.
        self.assertEqual(
            services._sanitize_filename_component('a/b:c*d?"e<f>g|h'), "a b c d e f g h"
        )

    def test_sanitize_filename_component_keeps_spaces_by_default(self):
        self.assertEqual(services._sanitize_filename_component("  Ana   Paula  "), "Ana Paula")

    def test_sanitize_filename_component_strip_mode_removes_spaces(self):
        self.assertEqual(
            services._sanitize_filename_component(
                "Ana Paula", space_mode=DocumentTemplate.FilenameSpaceMode.STRIP
            ),
            "AnaPaula",
        )

    def test_sanitize_filename_component_replace_mode_uses_custom_separator(self):
        self.assertEqual(
            services._sanitize_filename_component(
                "Ana Paula da Silva",
                space_mode=DocumentTemplate.FilenameSpaceMode.REPLACE,
                space_replacement=".",
            ),
            "Ana.Paula.da.Silva",
        )

    def test_sanitize_filename_component_replace_mode_falls_back_to_hyphen(self):
        self.assertEqual(
            services._sanitize_filename_component(
                "Ana Paula",
                space_mode=DocumentTemplate.FilenameSpaceMode.REPLACE,
                space_replacement="",
            ),
            "Ana-Paula",
        )

    def test_build_item_filename_keeps_spaces_by_default(self):
        template = self._build_template(filename_pattern="{1}_{2}")
        job = GenerationJob(template=template, name="Lote")
        name = services.build_item_filename(
            job, 1, {"coluna_1": "Ana", "coluna_2": "Empresa A"}, set()
        )
        self.assertEqual(name, "Ana_Empresa A.pdf")

    def test_build_item_filename_strip_mode(self):
        template = self._build_template(
            filename_pattern="{1}_{2}",
            filename_space_mode=DocumentTemplate.FilenameSpaceMode.STRIP,
        )
        job = GenerationJob(template=template, name="Lote")
        name = services.build_item_filename(
            job, 1, {"coluna_1": "Ana", "coluna_2": "Empresa A"}, set()
        )
        self.assertEqual(name, "Ana_EmpresaA.pdf")

    def test_build_item_filename_replace_mode(self):
        template = self._build_template(
            filename_pattern="{1}_{2}",
            filename_space_mode=DocumentTemplate.FilenameSpaceMode.REPLACE,
            filename_space_replacement=".",
        )
        job = GenerationJob(template=template, name="Lote")
        name = services.build_item_filename(
            job, 1, {"coluna_1": "Ana", "coluna_2": "Empresa A"}, set()
        )
        self.assertEqual(name, "Ana_Empresa.A.pdf")

    def test_build_item_filename_falls_back_without_pattern(self):
        template = self._build_template(filename_pattern="")
        job = GenerationJob(template=template, name="Meu Lote")
        name = services.build_item_filename(job, 3, {"coluna_1": "Ana"}, set())
        self.assertEqual(name, "meu_lote-linha-3.pdf")

    def test_build_item_filename_avoids_collision(self):
        template = self._build_template(filename_pattern="{1}")
        job = GenerationJob(template=template, name="Lote")
        used = set()
        first = services.build_item_filename(job, 1, {"coluna_1": "Ana"}, used)
        second = services.build_item_filename(job, 2, {"coluna_1": "Ana"}, used)
        self.assertEqual(first, "Ana.pdf")
        self.assertEqual(second, "Ana-2.pdf")

    def test_strip_accents_replaces_diacritics_with_plain_letters(self):
        self.assertEqual(services._strip_accents("Ana Cecília Ü"), "Ana Cecilia U")

    def test_build_item_filename_strips_accents_when_enabled(self):
        template = self._build_template(filename_pattern="{1}_{2}", filename_strip_accents=True)
        job = GenerationJob(template=template, name="Lote")
        name = services.build_item_filename(
            job, 1, {"coluna_1": "José Ção", "coluna_2": "São Paulo"}, set()
        )
        self.assertEqual(name, "Jose Cao_Sao Paulo.pdf")

    def test_build_item_filename_keeps_accents_by_default(self):
        template = self._build_template(filename_pattern="{1}")
        job = GenerationJob(template=template, name="Lote")
        name = services.build_item_filename(job, 1, {"coluna_1": "José"}, set())
        self.assertEqual(name, "José.pdf")

    def test_build_item_filename_uppercases_all_columns_without_selection(self):
        template = self._build_template(
            filename_pattern="{1}_{2}", filename_case=DocumentTemplate.FilenameCase.UPPER
        )
        job = GenerationJob(template=template, name="Lote")
        name = services.build_item_filename(
            job, 1, {"coluna_1": "ana", "coluna_2": "empresa a"}, set()
        )
        self.assertEqual(name, "ANA_EMPRESA A.pdf")

    def test_build_item_filename_lowercases_only_selected_column(self):
        template = self._build_template(
            filename_pattern="{1}-{2}",
            filename_case=DocumentTemplate.FilenameCase.LOWER,
            filename_case_columns=[2],
        )
        job = GenerationJob(template=template, name="Lote")
        name = services.build_item_filename(
            job, 1, {"coluna_1": "ANA", "coluna_2": "EMPRESA A"}, set()
        )
        # Só a coluna 2 (marcada) vira minúscula; a 1 fica como veio do Excel.
        self.assertEqual(name, "ANA-empresa a.pdf")

    def test_process_job_names_output_files_by_pattern(self):
        # display_filename é o nome como o usuário vê (download avulso e
        # ZIP); output_pdf.name é só o caminho interno no Storage, que o
        # Django sempre normaliza (troca espaço por "_") — não é o que este
        # teste quer verificar.
        template = self._build_template(filename_pattern="{1}_{2}")
        job = GenerationJob.objects.create(
            user=self.user,
            template=template,
            name="Job",
            source_excel=self._build_excel_upload([["Ana", "Empresa A"], ["Ana", "Empresa A"]]),
            kind=GenerationJob.Kind.PREVIEW,
            status=GenerationJob.Status.QUEUED,
        )
        process_job(job)
        job.refresh_from_db()
        names = sorted(item.display_filename for item in job.items.order_by("id"))
        # Linha 1 da planilha é o cabeçalho, então as duas linhas de dados são
        # numeradas 2 e 3 — o sufixo de desempate usa esse número de linha.
        self.assertEqual(names, ["Ana_Empresa A-3.pdf", "Ana_Empresa A.pdf"])


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT)
class WrapGrowDirectionTests(TestCase):
    """ "Quebrar linha" não podia se comportar como "Cortar": sem limite de
    linhas, e a caixa cresce para cima ou para baixo sem sobrescrever um campo
    vizinho — a base (modo "cima") ou o topo (modo "baixo") fica parado
    enquanto o outro lado se move conforme o texto ganha linhas."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="wrap-user", password="senha123")
        font_path = Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
        cls.font = FontAsset.objects.create(
            user=cls.user,
            name="Dejavu Sans",
            family="Dejavu Sans",
            variant="Regular",
            file=SimpleUploadedFile(
                "DejaVuSans.ttf", font_path.read_bytes(), content_type="font/ttf"
            ),
        )

    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)

    def test_fit_text_lines_does_not_cap_wrap_mode(self):
        lines, size = services._fit_text_lines(
            "uma frase razoavelmente longa para quebrar em varias linhas de texto",
            "Helvetica",
            18,
            120,
            TemplateField.OverflowMode.WRAP,
            1,
        )
        self.assertGreater(len(lines), 1)
        self.assertEqual(size, 18)

    def test_fit_text_lines_truncate_still_caps_at_max_lines(self):
        lines, _size = services._fit_text_lines(
            "uma frase razoavelmente longa para quebrar em varias linhas de texto",
            "Helvetica",
            18,
            120,
            TemplateField.OverflowMode.TRUNCATE,
            1,
        )
        self.assertEqual(len(lines), 1)

    def _build_excel_upload(self, rows) -> SimpleUploadedFile:
        temp_dir = Path(tempfile.mkdtemp(prefix="wrap-xlsx-"))
        xlsx_path = temp_dir / "dados.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Texto"])
        for row in rows:
            sheet.append(row)
        workbook.save(xlsx_path)
        content = xlsx_path.read_bytes()
        shutil.rmtree(temp_dir, ignore_errors=True)
        return SimpleUploadedFile(
            "dados.xlsx",
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _render(self, *, grow_direction, text):
        buffer = io.BytesIO()
        pdf_canvas = canvas.Canvas(buffer, pagesize=(400, 200))
        pdf_canvas.showPage()
        pdf_canvas.save()
        template = DocumentTemplate.objects.create(
            user=self.user,
            name="Wrap",
            slug=f"wrap-{grow_direction}-{len(text)}",
            background_pdf=SimpleUploadedFile(
                "background.pdf", buffer.getvalue(), content_type="application/pdf"
            ),
        )
        update_template_pdf_metadata(template)
        TemplateField.objects.create(
            template=template,
            name="texto",
            excel_column="1",
            x=20,
            y=100,
            width=80,
            height=24,
            font=self.font,
            font_size=16,
            line_height=1.2,
            order_index=1,
            overflow_mode=TemplateField.OverflowMode.WRAP,
            grow_direction=grow_direction,
        )
        job = GenerationJob.objects.create(
            user=self.user,
            template=template,
            name="Job",
            source_excel=self._build_excel_upload([[text]]),
            kind=GenerationJob.Kind.PREVIEW,
            status=GenerationJob.Status.QUEUED,
        )
        process_job(job)
        job.refresh_from_db()
        item = job.items.get()
        with item.output_pdf.open("rb") as output_file:
            bbox, _size = _ink_bbox(output_file.read(), resolution=150)
        self.assertIsNotNone(bbox, f"nada desenhado para {text!r} ({grow_direction})")
        return bbox

    # A mesma palavra repetida em todas as linhas (em vez de palavras
    # diferentes) mantém ascendentes/descendentes idênticos entre a versão
    # curta e a longa — senão a caixa de tinta varia com as letras da linha,
    # não só com a posição, e o teste vira falso positivo/negativo.
    _SHORT_TEXT = "um"
    _LONG_TEXT = "um " * 8

    def test_grow_down_keeps_top_fixed_and_extends_bottom(self):
        short_bbox = self._render(
            grow_direction=TemplateField.GrowDirection.DOWN, text=self._SHORT_TEXT
        )
        long_bbox = self._render(
            grow_direction=TemplateField.GrowDirection.DOWN, text=self._LONG_TEXT
        )
        self.assertAlmostEqual(short_bbox[1], long_bbox[1], delta=2)
        self.assertGreater(long_bbox[3], short_bbox[3] + 5)

    def test_grow_up_keeps_bottom_fixed_and_extends_top(self):
        short_bbox = self._render(
            grow_direction=TemplateField.GrowDirection.UP, text=self._SHORT_TEXT
        )
        long_bbox = self._render(
            grow_direction=TemplateField.GrowDirection.UP, text=self._LONG_TEXT
        )
        self.assertAlmostEqual(short_bbox[3], long_bbox[3], delta=2)
        self.assertLess(long_bbox[1], short_bbox[1] - 5)


def _dark_pixel_count(image) -> int:
    # Soma pixels abaixo de um limiar de cinza — não precisa ser exato:
    # o que os testes comparam é a diferença entre com/sem decoração.
    histogram = image.histogram()
    return sum(histogram[:200])


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT)
class TextDecorationTests(TestCase):
    """Sublinhado e tachado desenham uma linha vetorial extra usando a
    métrica real da fonte (post.underlinePosition/OS-2 yStrikeoutPosition),
    não um número fixo — por isso o teste é "apareceu mais tinta", não uma
    posição exata em pontos."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="decoration-user", password="senha123"
        )
        font_path = Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
        cls.font = FontAsset.objects.create(
            user=cls.user,
            name="Dejavu Sans",
            family="Dejavu Sans",
            variant="Regular",
            file=SimpleUploadedFile(
                "DejaVuSans.ttf", font_path.read_bytes(), content_type="font/ttf"
            ),
        )

    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)

    def _build_excel_upload(self, text) -> SimpleUploadedFile:
        temp_dir = Path(tempfile.mkdtemp(prefix="decoration-xlsx-"))
        xlsx_path = temp_dir / "dados.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Texto"])
        sheet.append([text])
        workbook.save(xlsx_path)
        content = xlsx_path.read_bytes()
        shutil.rmtree(temp_dir, ignore_errors=True)
        return SimpleUploadedFile(
            "dados.xlsx",
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _render(self, *, underline=False, strikethrough=False, text="A     B"):
        buffer = io.BytesIO()
        pdf_canvas = canvas.Canvas(buffer, pagesize=(400, 120))
        pdf_canvas.showPage()
        pdf_canvas.save()
        template = DocumentTemplate.objects.create(
            user=self.user,
            name="Decoração",
            slug=f"decoracao-{underline}-{strikethrough}",
            background_pdf=SimpleUploadedFile(
                "background.pdf", buffer.getvalue(), content_type="application/pdf"
            ),
        )
        update_template_pdf_metadata(template)
        TemplateField.objects.create(
            template=template,
            name="texto",
            excel_column="1",
            x=20,
            y=80,
            width=0,
            height=24,
            font=self.font,
            font_size=40,
            order_index=1,
            text_underline=underline,
            text_strikethrough=strikethrough,
        )
        job = GenerationJob.objects.create(
            user=self.user,
            template=template,
            name="Job",
            source_excel=self._build_excel_upload(text),
            kind=GenerationJob.Kind.PREVIEW,
            status=GenerationJob.Status.QUEUED,
        )
        process_job(job)
        job.refresh_from_db()
        item = job.items.get()
        with item.output_pdf.open("rb") as output_file:
            return _rasterize(output_file.read(), resolution=150)

    def test_underline_adds_ink(self):
        base = _dark_pixel_count(self._render())
        underlined = _dark_pixel_count(self._render(underline=True))
        self.assertGreater(underlined, base + 20)

    def test_strikethrough_adds_ink(self):
        base = _dark_pixel_count(self._render())
        struck = _dark_pixel_count(self._render(strikethrough=True))
        self.assertGreater(struck, base + 20)

    def test_both_together_add_more_ink_than_either_alone(self):
        underlined = _dark_pixel_count(self._render(underline=True))
        both = _dark_pixel_count(self._render(underline=True, strikethrough=True))
        self.assertGreater(both, underlined + 20)
