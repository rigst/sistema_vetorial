from django import forms
from openpyxl import load_workbook

from editor.models import DocumentTemplate

from .models import GenerationJob


class GenerationJobForm(forms.ModelForm):
    def clean_source_excel(self):
        uploaded = self.cleaned_data["source_excel"]
        suffix = uploaded.name.lower().rsplit(".", 1)[-1] if "." in uploaded.name else ""
        if suffix not in {"xlsx", "xlsm"}:
            raise forms.ValidationError("Envie um arquivo Excel .xlsx ou .xlsm.")
        uploaded.seek(0)
        try:
            workbook = load_workbook(uploaded, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True, max_row=2))
            workbook.close()
        except Exception as exc:
            uploaded.seek(0)
            raise forms.ValidationError(
                "Não foi possível ler o Excel. Verifique se o arquivo está íntegro e com formato válido."
            ) from exc
        finally:
            uploaded.seek(0)

        if not rows or not rows[0]:
            raise forms.ValidationError("O Excel precisa ter uma primeira linha de cabeçalho.")

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        if not any(headers):
            raise forms.ValidationError(
                "A primeira linha do Excel está vazia. Informe um cabeçalho válido."
            )
        self._uploaded_headers = headers
        return uploaded

    class Meta:
        model = GenerationJob
        fields = ["name", "template", "source_excel"]
        labels = {
            "name": "Nome do lote",
            "template": "Template",
            "source_excel": "Planilha Excel (.xlsx)",
        }
        widgets = {
            "name": forms.TextInput(attrs={"placeholder": "Ex.: Lote abril 2026"}),
            "source_excel": forms.FileInput(attrs={"accept": ".xlsx,.xlsm"}),
        }

    def __init__(self, *args, **kwargs):
        user = kwargs.pop("user")
        template_id = kwargs.pop("template_id", None)
        super().__init__(*args, **kwargs)
        self.fields["template"].queryset = DocumentTemplate.objects.filter(
            user=user, is_active=True
        ).order_by("name")
        if template_id:
            self.fields["template"].initial = template_id

    def clean(self):
        cleaned = super().clean()
        template = cleaned.get("template")
        headers = getattr(self, "_uploaded_headers", []) or []
        if template and headers:
            invalid_columns = []
            max_column = len(headers)
            for field in template.fields.all():
                if not field.excel_column or not str(field.excel_column).isdigit():
                    continue
                column_number = int(field.excel_column)
                if column_number < 1 or column_number > max_column:
                    invalid_columns.append(f"{field.name} -> coluna {column_number}")
            if invalid_columns:
                raise forms.ValidationError(
                    "O Excel não possui todas as colunas numéricas esperadas pelo template: "
                    + ", ".join(invalid_columns)
                )
        return cleaned
