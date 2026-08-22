from __future__ import annotations

import io
import math
import re
import unicodedata
import zipfile
from functools import lru_cache
from pathlib import Path
from tempfile import NamedTemporaryFile

import pikepdf
from django.core.files.base import ContentFile
from fontTools.pens.basePen import BasePen
from fontTools.ttLib import TTFont as FontToolsTTFont
from openpyxl import load_workbook
from pikepdf import Name, Page, Pdf
from reportlab.lib.colors import HexColor
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.pdfgen.canvas import FILL_NON_ZERO

from editor.models import DocumentTemplate, TemplateField
from editor.services import PageGeometry, read_page_geometry

from .models import GenerationItem, GenerationJob

# Sintaxe compartilhada por dois lugares: a coluna de origem de um campo
# ("{3}-{5}" junta a coluna 3 e a 5 com um traço no meio) e o nome do
# arquivo gerado (DocumentTemplate.filename_pattern, ex.: "{1}_{2}"). Um
# número solto sem chaves ("3") continua valendo — é o formato antigo, uma
# coluna só, sem conector.
COLUMN_REF_RE = re.compile(r"\{(\d+)\}")


def extract_column_refs(pattern: str) -> list[int]:
    """Todo número de coluna citado num padrão — usado para validar que o
    Excel enviado tem colunas suficientes antes de gerar qualquer arquivo."""
    if not pattern:
        return []
    pattern = str(pattern).strip()
    if pattern.isdigit():
        return [int(pattern)]
    return [int(match) for match in COLUMN_REF_RE.findall(pattern)]


def render_column_template(pattern: str, payload: dict, column_transform=None) -> str:
    """Resolve um padrão de colunas contra os valores de uma linha do Excel.

    "{3}-{5}" com coluna 3 = "Recife" e coluna 5 = "PE" vira "Recife-PE": o
    texto entre chaves é copiado como está (o conector), só o que está
    dentro de `{}` é trocado pelo valor da coluna. Um número solto ("3")
    continua funcionando como antes — uma coluna só, sem chaves.

    `column_transform`, quando informado, recebe (numero_da_coluna, valor) e
    devolve o valor a usar — permite aplicar a Transformação de texto só nas
    colunas escolhidas antes de juntá-las (ver `resolve_field_raw_value`).
    """
    if not pattern:
        return ""
    pattern = str(pattern).strip()

    def _value_for(column_number: int) -> str:
        value = str(payload.get(f"coluna_{column_number}", "") or "")
        if column_transform:
            value = column_transform(column_number, value)
        return value

    if pattern.isdigit():
        return _value_for(int(pattern))

    def _substitute(match: re.Match) -> str:
        return _value_for(int(match.group(1)))

    return COLUMN_REF_RE.sub(_substitute, pattern)


def _humanize_generation_error(exc: Exception) -> str:
    if isinstance(exc, ValueError):
        return str(exc)
    if isinstance(exc, pikepdf.PdfError):
        return "Não foi possível compor o PDF final desta linha com o modelo enviado."
    if isinstance(exc, OSError):
        return "Não foi possível acessar um dos arquivos necessários para gerar esta linha."
    return "Ocorreu um erro inesperado ao gerar esta linha do PDF."


def _register_font(field: TemplateField) -> str:
    font_name = f"font_{field.font_id}"
    registered = pdfmetrics.getRegisteredFontNames()
    if font_name not in registered:
        pdfmetrics.registerFont(TTFont(font_name, field.font.file.path))
    return font_name


def _normalize_value(value) -> str:
    if value is None:
        return ""
    return unicodedata.normalize("NFC", str(value)).strip()


DEFAULT_TITLE_EXCEPTIONS = {
    "a",
    "as",
    "o",
    "os",
    "de",
    "da",
    "das",
    "do",
    "dos",
    "e",
    "em",
    "com",
    "para",
    "por",
    "na",
    "nas",
    "no",
    "nos",
    "à",
    "às",
}


def _normalize_unicode_text(text: str) -> str:
    return unicodedata.normalize("NFC", text or "")


def _title_word_pt_br(word: str) -> str:
    if not word:
        return word
    first = word[:1].upper()
    rest = word[1:]
    return first + rest


def _smart_title_token(token: str, exceptions: set[str], is_edge: bool) -> str:
    parts = re.split(r"([-/'’])", token)
    normalized_parts = []
    for part in parts:
        if part in {"-", "/", "'", "’"}:
            normalized_parts.append(part)
            continue
        lowered = part.lower()
        if not is_edge and lowered in exceptions:
            normalized_parts.append(lowered)
        else:
            normalized_parts.append(_title_word_pt_br(lowered))
    return "".join(normalized_parts)


def _smart_title(text: str, exceptions_raw: str) -> str:
    text = _normalize_unicode_text(text)
    words = text.split()
    if not words:
        return ""
    exceptions = {
        _normalize_unicode_text(item.strip()).lower()
        for item in (exceptions_raw or "").split(",")
        if item.strip()
    } or DEFAULT_TITLE_EXCEPTIONS
    return " ".join(
        _smart_title_token(word, exceptions, index in {0, len(words) - 1})
        for index, word in enumerate(words)
    )


def _load_font_cmap(font_path: str) -> set[int]:
    font = FontToolsTTFont(font_path)
    try:
        cmap = set()
        for table in font["cmap"].tables:
            cmap.update(table.cmap.keys())
        return cmap
    finally:
        font.close()


@lru_cache(maxsize=256)
def _cached_font_cmap(font_path: str) -> frozenset[int]:
    return frozenset(_load_font_cmap(font_path))


def get_missing_font_chars(field: TemplateField, text: str) -> list[str]:
    text = _normalize_unicode_text(text)
    if not text:
        return []
    supported_codepoints = _cached_font_cmap(field.font.file.path)
    return sorted(
        {char for char in text if not char.isspace() and ord(char) not in supported_codepoints}
    )


def validate_font_supports_text(field: TemplateField, text: str) -> None:
    text = _normalize_unicode_text(text)
    if not text:
        return
    missing_chars = get_missing_font_chars(field, text)
    if missing_chars:
        chars_display = ", ".join(repr(char) for char in missing_chars[:10])
        raise ValueError(
            f"A fonte '{field.font.name}' nao suporta todos os caracteres exigidos pelo texto. Faltando: {chars_display}."
        )


def _apply_text_transform(field: TemplateField, value: str) -> str:
    if field.text_transform == TemplateField.TextTransform.LOWER:
        return value.lower()
    if field.text_transform == TemplateField.TextTransform.UPPER:
        return value.upper()
    if field.text_transform == TemplateField.TextTransform.TITLE:
        return " ".join(_title_word_pt_br(token.lower()) for token in value.split())
    if field.text_transform == TemplateField.TextTransform.TITLE_SMART:
        return _smart_title(value, field.transform_exceptions)
    return value


def resolve_field_raw_value(field: TemplateField, payload: dict) -> tuple[str, bool]:
    """Junta as colunas do Excel referenciadas em `field.excel_column`.

    Quando o campo junta mais de uma coluna e tem colunas escolhidas em
    `transform_columns`, a Transformação (text_transform) é aplicada coluna a
    coluna, antes de juntar — só nas colunas marcadas, deixando o conector e
    as demais colunas como vieram do Excel. Sem essa escolha (lista vazia, o
    padrão), o comportamento é o de sempre: a transformação roda depois, em
    `format_field_value`, sobre o texto já unido.

    Retorna (valor_bruto, transformacao_ja_aplicada).
    """
    column_refs = extract_column_refs(field.excel_column)
    selected_columns = set(field.transform_columns or [])
    if len(column_refs) > 1 and selected_columns:

        def _column_transform(column_number: int, value: str) -> str:
            if column_number in selected_columns:
                return _apply_text_transform(field, value)
            return value

        return render_column_template(field.excel_column, payload, _column_transform), True
    return render_column_template(field.excel_column, payload), False


def format_field_value(
    field: TemplateField, raw_value: str, *, transform_applied: bool = False
) -> str:
    value = _normalize_unicode_text("" if raw_value is None else str(raw_value))
    if field.trim_whitespace:
        value = value.strip()

    if not value:
        value = field.empty_value or ""

    if field.value_type == TemplateField.ValueType.INTEGER and value:
        stripped = value.strip()
        negative = stripped.startswith("-")
        digits = re.sub(r"\D", "", stripped)
        if digits:
            number = str(int(digits))
            if field.integer_min_digits:
                number = number.zfill(field.integer_min_digits)
            if negative and field.integer_keep_sign:
                value = f"-{number}"
            else:
                value = number

    if not transform_applied:
        value = _apply_text_transform(field, value)

    return _normalize_unicode_text(value)


def load_excel_rows(excel_path: str):
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return [], []
    headers = [_normalize_value(value) for value in rows[0]]
    # A largura "usada" que o Excel relata costuma sobrar além do último
    # cabeçalho de verdade (células vazias que já foram formatadas ou
    # tocadas alguma vez) — sem isto, "Cabeçalhos detectados" na tela do job
    # exibia um monte de vírgulas soltas depois dos nomes reais.
    while headers and not headers[-1]:
        headers.pop()
    data_rows = []
    for row_index, row in enumerate(rows[1:], start=2):
        payload = {}
        has_value = False
        for idx, header in enumerate(headers):
            value = _normalize_value(row[idx] if idx < len(row) else "")
            payload[f"coluna_{idx + 1}"] = value
            if header:
                payload[header] = value
            has_value = has_value or bool(value)
        if has_value:
            data_rows.append((row_index, payload))
    return headers, data_rows


def get_template_expected_headers(job: GenerationJob) -> list[str]:
    return [
        f"Coluna {column_number}"
        for field in job.template.fields.all()
        for column_number in extract_column_refs(field.excel_column)
    ]


def _fit_text_lines(
    text: str, font_name: str, font_size: float, max_width: float, mode: str, max_lines: int
):
    if not text:
        return [""], font_size

    # "Quebrar linha" existe para mostrar o texto inteiro — limitar a
    # max_lines aqui derrubaria o resto do texto e o efeito seria idêntico a
    # "Cortar". max_lines só governa este modo indiretamente, via a caixa
    # ficar maior (grow_direction, calculado depois em _draw_field).
    line_cap = None if mode == TemplateField.OverflowMode.WRAP else max_lines

    if max_width <= 0:
        return text.splitlines()[:line_cap], font_size

    def wrap(current_size: float):
        words = text.split()
        if not words:
            return [""]
        lines = []
        current = words[0]
        for word in words[1:]:
            candidate = f"{current} {word}"
            if pdfmetrics.stringWidth(candidate, font_name, current_size) <= max_width:
                current = candidate
            else:
                lines.append(current)
                current = word
        lines.append(current)
        return lines

    if mode == TemplateField.OverflowMode.WRAP:
        return wrap(font_size), font_size

    if mode == TemplateField.OverflowMode.SHRINK:
        size = font_size
        while size >= 4:
            lines = wrap(size)
            if (
                len(lines) <= max_lines
                and max(
                    pdfmetrics.stringWidth(line, font_name, size)
                    for line in lines
                    if line is not None
                )
                <= max_width
            ):
                return lines, size
            size -= 0.5
        return wrap(max(4, size)), max(4, size)

    if mode == TemplateField.OverflowMode.ERROR:
        if pdfmetrics.stringWidth(text, font_name, font_size) > max_width:
            raise ValueError("Texto excede a largura configurada para o campo.")
        return [text], font_size

    trimmed = text
    while trimmed and pdfmetrics.stringWidth(trimmed, font_name, font_size) > max_width:
        trimmed = trimmed[:-1]
    return [trimmed], font_size


# Constantes de métrica do Fabric.js (editor visual): a primeira baseline fica a
# fontSize * 1.13 * (1 - 0.222) do topo da caixa. Usar a mesma fórmula aqui faz o
# PDF gerado coincidir com o que o editor mostra na tela.
FABRIC_FONT_SIZE_MULT = 1.13
FABRIC_FONT_SIZE_FRACTION = 0.222
FIRST_BASELINE_FACTOR = FABRIC_FONT_SIZE_MULT * (1 - FABRIC_FONT_SIZE_FRACTION)


class _GlyphOutlinePen(BasePen):
    """Traduz o contorno de um glifo para um path do reportlab.

    A BasePen já converte as curvas quadráticas do TrueType em cúbicas — que é
    a única forma de curva que o PDF conhece — e resolve glifos compostos.
    """

    def __init__(self, glyph_set, path, scale: float, origin_x: float, origin_y: float):
        super().__init__(glyph_set)
        self._path = path
        self._scale = scale
        self._origin_x = origin_x
        self._origin_y = origin_y

    def _place(self, point):
        return (
            self._origin_x + point[0] * self._scale,
            self._origin_y + point[1] * self._scale,
        )

    def _moveTo(self, point):
        self._path.moveTo(*self._place(point))

    def _lineTo(self, point):
        self._path.lineTo(*self._place(point))

    def _curveToOne(self, point1, point2, point3):
        self._path.curveTo(*self._place(point1), *self._place(point2), *self._place(point3))

    def _closePath(self):
        self._path.close()

    def _endPath(self):
        pass


@lru_cache(maxsize=32)
def _decoration_metrics(font_path: str) -> dict:
    """Posição/espessura do sublinhado (tabela post) e do tachado (OS/2),
    como fração de font_size — lidas do próprio desenho da fonte em vez de
    um número fixo que só ficaria bem numa fonte específica."""
    ttfont = FontToolsTTFont(font_path)
    units_per_em = float(ttfont["head"].unitsPerEm) or 1000.0
    post = ttfont["post"]
    os2 = ttfont.get("OS/2")
    underline_position = float(getattr(post, "underlinePosition", None) or -units_per_em * 0.1)
    underline_thickness = float(getattr(post, "underlineThickness", None) or units_per_em * 0.05)
    if os2 is not None:
        strikeout_position = float(getattr(os2, "yStrikeoutPosition", None) or units_per_em * 0.3)
        strikeout_thickness = float(getattr(os2, "yStrikeoutSize", None) or units_per_em * 0.05)
    else:
        strikeout_position = units_per_em * 0.3
        strikeout_thickness = units_per_em * 0.05
    return {
        "underline_position": underline_position / units_per_em,
        "underline_thickness": underline_thickness / units_per_em,
        "strikeout_position": strikeout_position / units_per_em,
        "strikeout_thickness": strikeout_thickness / units_per_em,
    }


@lru_cache(maxsize=32)
def _outline_font(font_path: str):
    """Fonte aberta pelo fontTools, pronta para extrair contornos.

    O cmap é a união de todas as tabelas, igual ao usado na validação de
    caracteres, para que validar e desenhar concordem sobre o que a fonte cobre.
    """
    ttfont = FontToolsTTFont(font_path)
    codepoints = {}
    for table in ttfont["cmap"].tables:
        codepoints.update(table.cmap)
    return ttfont.getGlyphSet(), codepoints, float(ttfont["head"].unitsPerEm)


def _text_outline_path(pdf_canvas, font_path: str, font_size: float, x: float, y: float, text: str):
    """Monta o path com os contornos de cada letra da linha.

    O texto vira geometria: o PDF de saída não carrega fonte embutida nem
    operadores de texto, só curvas — que é o que um fluxo vetorial (corte,
    plotter, gravação) precisa receber.
    """
    glyph_set, codepoints, units_per_em = _outline_font(font_path)
    scale = font_size / units_per_em
    path = pdf_canvas.beginPath()
    pen_x = 0.0
    drew_any = False
    for char in text:
        glyph_name = codepoints.get(ord(char), ".notdef")
        glyph = glyph_set.get(glyph_name)
        if glyph is None:
            continue
        pen = _GlyphOutlinePen(glyph_set, path, scale, x + pen_x, y)
        glyph.draw(pen)
        # O avanço vem do hmtx, a mesma fonte de medida do stringWidth do
        # reportlab: a linha desenhada tem exatamente a largura já calculada.
        pen_x += glyph.width * scale
        drew_any = True
    return path if drew_any else None


def _draw_field(pdf_canvas, field: TemplateField, value: str, page_height: float):
    # A fonte é registrada só para medir (stringWidth); o desenho sai em curvas,
    # então nenhum recurso de fonte chega ao PDF final.
    font_name = _register_font(field)
    font_path = field.font.file.path
    validate_font_supports_text(field, value)
    font_size = float(field.font_size)
    max_width = float(field.width)
    max_lines = max(field.max_lines, 1)
    lines, adjusted_size = _fit_text_lines(
        value,
        font_name,
        font_size,
        max_width,
        field.overflow_mode,
        max_lines,
    )

    line_height = max(adjusted_size * float(field.line_height), adjusted_size * 0.2)
    rotation = float(field.rotation or 0)

    def aligned_x(draw_x, line):
        if field.text_align == TemplateField.TextAlign.CENTER and max_width > 0:
            return draw_x + (max_width - pdfmetrics.stringWidth(line, font_name, adjusted_size)) / 2
        if field.text_align == TemplateField.TextAlign.RIGHT and max_width > 0:
            return draw_x + max_width - pdfmetrics.stringWidth(line, font_name, adjusted_size)
        return draw_x

    def baseline_y(index):
        offset = -((FIRST_BASELINE_FACTOR * adjusted_size) + (index * line_height))
        if field.grow_direction == TemplateField.GrowDirection.UP:
            # A última linha fica onde a única linha ficaria numa caixa que
            # nunca precisou crescer (a "base" original) — linhas extras
            # empurram o bloco para cima, sem inverter a ordem de leitura:
            # a primeira linha do texto continua sendo a de cima.
            offset += (len(lines) - 1) * line_height
        return offset

    def draw_text_line(
        draw_x, baseline_y, line, *, fill_color=None, stroke_color=None, line_width=1, alpha=1
    ):
        if not line:
            return
        path = _text_outline_path(pdf_canvas, font_path, adjusted_size, draw_x, baseline_y, line)
        if path is None:
            return
        pdf_canvas.saveState()
        pdf_canvas.setLineWidth(line_width)
        # Junções e pontas arredondadas: o contorno acompanha a curva da letra
        # em vez de criar bicos nos vértices.
        pdf_canvas.setLineJoin(1)
        pdf_canvas.setLineCap(1)
        if stroke_color is not None:
            pdf_canvas.setStrokeColor(stroke_color)
            pdf_canvas.setStrokeAlpha(alpha)
        if fill_color is not None:
            pdf_canvas.setFillColor(fill_color)
            # A letra em si é opaca; as cópias do halo de blur é que usam alfa.
            pdf_canvas.setFillAlpha(1 if stroke_color is not None else alpha)
        pdf_canvas.drawPath(
            path,
            stroke=1 if stroke_color is not None else 0,
            fill=1 if fill_color is not None else 0,
            # Regra não-zero: é a que o TrueType assume para vazar contra-formas
            # (o miolo do "o") sem furar contornos sobrepostos.
            fillMode=FILL_NON_ZERO,
        )
        pdf_canvas.restoreState()

    def draw_decoration(draw_x, line_baseline, line, *, fill_color):
        if not line or not (field.text_underline or field.text_strikethrough):
            return
        width = pdfmetrics.stringWidth(line, font_name, adjusted_size)
        if width <= 0:
            return
        metrics = _decoration_metrics(font_path)
        pdf_canvas.saveState()
        pdf_canvas.setFillColor(fill_color)
        if field.text_underline:
            y = line_baseline + metrics["underline_position"] * adjusted_size
            thickness = max(metrics["underline_thickness"] * adjusted_size, 0.4)
            pdf_canvas.rect(draw_x, y - thickness / 2, width, thickness, fill=1, stroke=0)
        if field.text_strikethrough:
            y = line_baseline + metrics["strikeout_position"] * adjusted_size
            thickness = max(metrics["strikeout_thickness"] * adjusted_size, 0.4)
            pdf_canvas.rect(draw_x, y - thickness / 2, width, thickness, fill=1, stroke=0)
        pdf_canvas.restoreState()

    pdf_canvas.saveState()
    # Origem local no canto superior-esquerdo da caixa; y cresce para baixo no
    # editor, então as baselines ficam em y locais negativos.
    pdf_canvas.translate(float(field.x), page_height - float(field.y))
    if rotation:
        pdf_canvas.rotate(-rotation)

    if field.border_enabled and float(field.border_size_ratio or 0) > 0:
        border_width = max(adjusted_size * float(field.border_size_ratio), 0.2)
        blur_factor = max(float(field.border_blur or 0), 0)
        border_alpha = min(max(float(field.border_opacity or 1), 0), 1)
        border_color = HexColor(field.border_color or "#000000")
        for index, line in enumerate(lines):
            line_baseline = baseline_y(index)
            x = aligned_x(0, line)
            if blur_factor > 0:
                shadow_samples = max(6 + int(blur_factor * 4), 6)
                shadow_radius = blur_factor * 0.35
                for sample in range(shadow_samples):
                    angle = (sample / shadow_samples) * math.tau
                    draw_text_line(
                        x + (shadow_radius * math.cos(angle)),
                        line_baseline + (shadow_radius * math.sin(angle)),
                        line,
                        fill_color=border_color,
                        alpha=max(border_alpha * 0.22, 0.05),
                    )
            draw_text_line(
                x,
                line_baseline,
                line,
                fill_color=HexColor(field.color or "#000000"),
                stroke_color=border_color,
                line_width=border_width,
                alpha=border_alpha,
            )
            draw_decoration(x, line_baseline, line, fill_color=HexColor(field.color or "#000000"))
        pdf_canvas.restoreState()
        return

    fill_color = HexColor(field.color or "#000000")
    for index, line in enumerate(lines):
        x = aligned_x(0, line)
        line_baseline = baseline_y(index)
        draw_text_line(x, line_baseline, line, fill_color=fill_color)
        draw_decoration(x, line_baseline, line, fill_color=fill_color)
    pdf_canvas.restoreState()


def _apply_page_rotation(pdf_canvas, geometry: PageGeometry) -> None:
    """Leva o espaço visível (o que o editor mostra) para o espaço da página.

    O /Rotate gira a página na hora de exibir, mas não mexe nas coordenadas do
    conteúdo. Sem esta compensação, um fundo girado receberia o texto deitado.
    """
    if geometry.rotation == 90:
        pdf_canvas.translate(geometry.width, 0)
        pdf_canvas.rotate(90)
    elif geometry.rotation == 180:
        pdf_canvas.translate(geometry.width, geometry.height)
        pdf_canvas.rotate(180)
    elif geometry.rotation == 270:
        pdf_canvas.translate(0, geometry.height)
        pdf_canvas.rotate(-90)


def _build_overlay_pdf(job: GenerationJob, payload: dict, geometry: PageGeometry) -> bytes:
    buffer = io.BytesIO()
    fields = list(job.template.fields.select_related("font").order_by("order_index", "id"))

    # O reportlab abre toda página com um "BT /F1 12 Tf ET" de cortesia, que
    # arrastaria um recurso de fonte para dentro do PDF vetorial. Ele pula esse
    # preâmbulo quando a fonte inicial é uma TTF dinâmica, então basta apontar
    # a primeira fonte do projeto.
    initial_font = None
    for field in fields:
        initial_font = _register_font(field)

    # A página do overlay nasce do tamanho exato da caixa do fundo: assim a
    # sobreposição acontece em escala 1:1 e o texto não é reamostrado.
    pdf_canvas = canvas.Canvas(
        buffer,
        pagesize=(geometry.width, geometry.height),
        initialFontName=initial_font,
    )
    pdf_canvas.saveState()
    _apply_page_rotation(pdf_canvas, geometry)

    for field in fields:
        raw_value, transform_applied = resolve_field_raw_value(field, payload)
        if not raw_value:
            raw_value = payload.get(field.name) or field.empty_value or field.name
            transform_applied = False
        value = format_field_value(field, raw_value, transform_applied=transform_applied)
        _draw_field(pdf_canvas, field, value, geometry.visible_height)
    pdf_canvas.restoreState()
    pdf_canvas.showPage()
    pdf_canvas.save()
    return buffer.getvalue()


def _merge_overlay(background_pdf_path: str, overlay_bytes: bytes, geometry: PageGeometry) -> bytes:
    """Sobrepõe o texto ao PDF de fundo sem tocar no fundo.

    O fundo é aberto e salvo pelo pikepdf, que copia os streams como estão: as
    imagens continuam com os mesmos bytes, resolução e compressão do arquivo
    enviado, e as caixas da página (MediaBox/CropBox) ficam intactas.
    """
    with NamedTemporaryFile(suffix=".pdf") as overlay_temp:
        overlay_temp.write(overlay_bytes)
        overlay_temp.flush()

        with Pdf.open(background_pdf_path) as base_pdf, Pdf.open(overlay_temp.name) as overlay_pdf:
            destination_page = Page(base_pdf.pages[0])
            overlay_page = Page(overlay_pdf.pages[0])
            overlay_box = [float(value) for value in overlay_page.obj.MediaBox]
            overlay_width = overlay_box[2] - overlay_box[0]
            overlay_height = overlay_box[3] - overlay_box[1]
            if (
                abs(overlay_width - geometry.width) > 0.01
                or abs(overlay_height - geometry.height) > 0.01
            ):
                raise ValueError(
                    "O tamanho da página do fundo mudou desde a última leitura do template. "
                    "Reenvie o fundo do projeto para atualizar as medidas."
                )

            # A colocação é feita à mão em vez de por add_overlay: o cálculo do
            # qpdf reencaixa o overlay quando a página tem /Rotate ou quando as
            # medidas não batem, e qualquer reencaixe significaria escalar o
            # texto. Aqui a matriz é sempre uma translação pura — escala 1:1.
            form = overlay_page.as_form_xobject()
            name = destination_page.add_resource(form, Name.XObject)
            destination_page.contents_add(b"q\n", prepend=True)
            destination_page.contents_add(
                f"Q\nq\n1 0 0 1 {geometry.x0:.4f} {geometry.y0:.4f} cm\n{name} Do\nQ\n".encode(),
                prepend=False,
            )
            destination_page.contents_coalesce()

            output = io.BytesIO()
            base_pdf.save(output)
            return output.getvalue()


_UNSAFE_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|\r\n\t]+')


def _strip_accents(value: str) -> str:
    """Troca à/ã/ç/ü... pela letra latina simples equivalente (a, a, c, u...).

    NFD separa cada letra acentuada em base + marca de acento (combining
    mark); descartando as marcas sobra só a base.
    """
    decomposed = unicodedata.normalize("NFD", value)
    return "".join(char for char in decomposed if not unicodedata.combining(char))


def _filename_column_transform(template: DocumentTemplate):
    """Fábrica do `column_transform` de `render_column_template` para o nome
    do arquivo: aplica remoção de acentos e maiúsculas/minúsculas só ao texto
    vindo das colunas do Excel — o conector literal do padrão não passa por
    aqui. A caixa respeita `filename_case_columns` do mesmo jeito que
    `TemplateField.transform_columns`: lista vazia = todas as colunas.
    """
    selected_columns = set(template.filename_case_columns or [])

    def _transform(column_number: int, value: str) -> str:
        if template.filename_strip_accents:
            value = _strip_accents(value)
        if template.filename_case != DocumentTemplate.FilenameCase.NONE and (
            not selected_columns or column_number in selected_columns
        ):
            value = (
                value.upper()
                if template.filename_case == DocumentTemplate.FilenameCase.UPPER
                else value.lower()
            )
        return value

    return _transform


def _sanitize_filename_component(
    value: str,
    *,
    space_mode: str = DocumentTemplate.FilenameSpaceMode.KEEP,
    space_replacement: str = "-",
) -> str:
    value = unicodedata.normalize("NFC", value or "").strip()
    # Caracteres proibidos em nome de arquivo viram espaço; o que fazer com
    # espaço (inclusive os que já vinham do texto) é decidido só depois, pelo
    # space_mode escolhido no template — por padrão o app não mexe neles.
    value = _UNSAFE_FILENAME_CHARS.sub(" ", value)
    value = re.sub(r"\s+", " ", value).strip()
    if space_mode == DocumentTemplate.FilenameSpaceMode.STRIP:
        value = value.replace(" ", "")
    elif space_mode == DocumentTemplate.FilenameSpaceMode.REPLACE:
        value = value.replace(" ", (space_replacement or "-").strip() or "-")
    return value


def build_item_filename(
    job: GenerationJob, row_number: int, payload: dict, used_names: set[str]
) -> str:
    """Nome do PDF de uma linha: o padrão do template (com variáveis {1},
    {2}...) quando configurado, senão o nome antigo (nome do job + linha).

    `used_names` acompanha o que já foi usado NESTE job — um padrão como
    "{1}" gera o mesmo nome para duas pessoas com o mesmo valor na coluna 1,
    e sem desempate uma sobrescreveria a outra dentro do ZIP.
    """
    pattern = job.template.filename_pattern
    base = (
        _sanitize_filename_component(
            render_column_template(pattern, payload, _filename_column_transform(job.template)),
            space_mode=job.template.filename_space_mode,
            space_replacement=job.template.filename_space_replacement,
        )
        if pattern
        else ""
    )
    if not base:
        base = f"{job.name.lower().replace(' ', '_')}-linha-{row_number}"

    filename = f"{base}.pdf"
    if filename.lower() in used_names:
        filename = f"{base}-{row_number}.pdf"
    used_names.add(filename.lower())
    return filename


def _create_item_pdf(
    job: GenerationJob,
    row_number: int,
    payload: dict,
    geometry: PageGeometry,
    used_filenames: set[str],
) -> GenerationItem:
    item = GenerationItem.objects.create(
        job=job, row_number=row_number, payload=payload, status=GenerationItem.Status.PROCESSING
    )
    try:
        overlay_bytes = _build_overlay_pdf(job, payload, geometry)
        output_bytes = _merge_overlay(job.template.background_pdf.path, overlay_bytes, geometry)
        filename = build_item_filename(job, row_number, payload, used_filenames)
        # O Storage do Django troca espaço por "_" no nome salvo em disco —
        # display_filename guarda o nome como o padrão realmente pediu, para
        # o download avulso e o ZIP mostrarem exatamente isso.
        item.display_filename = filename
        item.output_pdf.save(filename, ContentFile(output_bytes), save=False)
        item.status = GenerationItem.Status.COMPLETED
        item.error_message = ""
    except Exception as exc:
        item.status = GenerationItem.Status.FAILED
        item.error_message = _humanize_generation_error(exc)
    item.save()
    return item


def _build_zip_for_job(job: GenerationJob) -> None:
    completed_items = job.items.filter(status=GenerationItem.Status.COMPLETED, output_pdf__gt="")
    if not completed_items.exists():
        return

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for item in completed_items:
            with item.output_pdf.open("rb") as generated:
                entry_name = item.display_filename or Path(item.output_pdf.name).name
                archive.writestr(entry_name, generated.read())
    job.zip_file.save(
        f"{job.name.lower().replace(' ', '_')}.zip", ContentFile(buffer.getvalue()), save=False
    )


# Sem transação envolvendo o job inteiro: cada save de progresso precisa ficar
# visível para a conexão que responde o polling de status (jobs:status) — dentro
# de uma transaction.atomic o card de geração ficaria parado até o commit final.
def process_job(job: GenerationJob) -> GenerationJob:
    try:
        headers, data_rows = load_excel_rows(job.source_excel.path)
    except Exception:
        job.status = GenerationJob.Status.FAILED
        job.last_error = "Falha ao importar o Excel. Verifique se o arquivo está íntegro, com cabeçalho na primeira linha e formato válido."
        job.save(update_fields=["status", "last_error", "updated_at"])
        return job

    if not job.template.fields.exists():
        job.status = GenerationJob.Status.FAILED
        job.last_error = "O template precisa ter ao menos um campo configurado."
        job.save(update_fields=["status", "last_error", "updated_at"])
        return job

    expected_headers = get_template_expected_headers(job)
    max_column = len(headers)
    missing_headers = []
    for field in job.template.fields.all():
        for column_number in extract_column_refs(field.excel_column):
            if column_number > max_column:
                missing_headers.append(f"Coluna {column_number}")

    limit = 3 if job.kind == GenerationJob.Kind.PREVIEW else len(data_rows)
    selected_rows = data_rows[:limit]

    if not data_rows:
        job.status = GenerationJob.Status.FAILED
        job.last_error = "O Excel não possui linhas preenchidas após o cabeçalho."
        job.column_map = {
            "headers": headers,
            "expected_headers": expected_headers,
            "missing_headers": missing_headers,
            "sample_row_numbers": [],
        }
        job.save(update_fields=["status", "last_error", "column_map", "updated_at"])
        return job

    if missing_headers:
        job.status = GenerationJob.Status.FAILED
        job.last_error = (
            "O Excel não possui todos os cabeçalhos esperados pelo template. Faltando: "
            + ", ".join(missing_headers)
        )
        job.column_map = {
            "headers": headers,
            "expected_headers": expected_headers,
            "missing_headers": missing_headers,
            "sample_row_numbers": [row_number for row_number, _ in selected_rows],
        }
        job.save(update_fields=["status", "last_error", "column_map", "updated_at"])
        return job

    job.status = GenerationJob.Status.PROCESSING
    job.total_rows = len(selected_rows)
    job.processed_rows = 0
    job.success_rows = 0
    job.failed_rows = 0
    job.last_error = ""
    job.column_map = {
        "headers": headers,
        "expected_headers": expected_headers,
        "missing_headers": missing_headers,
        "sample_row_numbers": [row_number for row_number, _ in selected_rows],
    }
    job.save(
        update_fields=[
            "status",
            "total_rows",
            "processed_rows",
            "success_rows",
            "failed_rows",
            "last_error",
            "column_map",
            "updated_at",
        ]
    )
    job.items.all().delete()

    geometry = read_page_geometry(job.template.background_pdf.path)
    used_filenames: set[str] = set()
    for row_number, payload in selected_rows:
        item = _create_item_pdf(job, row_number, payload, geometry, used_filenames)
        job.processed_rows += 1
        if item.status == GenerationItem.Status.COMPLETED:
            job.success_rows += 1
        else:
            job.failed_rows += 1
            job.last_error = item.error_message
        job.save(
            update_fields=[
                "processed_rows",
                "success_rows",
                "failed_rows",
                "last_error",
                "updated_at",
            ]
        )

    if job.kind == GenerationJob.Kind.FULL:
        try:
            _build_zip_for_job(job)
        except Exception:
            job.status = GenerationJob.Status.FAILED
            job.last_error = "Falha ao exportar o ZIP final do job."
            job.save(update_fields=["status", "last_error", "updated_at"])
            return job

    job.status = (
        GenerationJob.Status.COMPLETED if job.failed_rows == 0 else GenerationJob.Status.FAILED
    )
    job.save(update_fields=["status", "zip_file", "updated_at"])
    return job


def clone_preview_to_full(preview_job: GenerationJob) -> GenerationJob:
    with preview_job.source_excel.open("rb") as source_file:
        full_job = GenerationJob.objects.create(
            user=preview_job.user,
            template=preview_job.template,
            name=f"{preview_job.name} - lote completo",
            source_excel=ContentFile(
                source_file.read(), name=Path(preview_job.source_excel.name).name
            ),
            kind=GenerationJob.Kind.FULL,
            status=GenerationJob.Status.QUEUED,
        )
    return process_job(full_job)
